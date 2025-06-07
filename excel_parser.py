# excel_parser.py
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from google.generativeai import types
import base64

def get_image_format(image_data):
    """Identifies image format from magic bytes. Defaults to 'png'."""
    if image_data.startswith(b'\xff\xd8\xff'):
        return 'jpeg'
    if image_data.startswith(b'\x89PNG\r\n\x1a\n'):
        return 'png'
    return 'png' # Default fallback

def parse_excel_to_gemini_parts(xlsx_path: str) -> list:
    """
    Parses an Excel file into a list of Parts suitable for the Gemini API.
    
    Returns:
        A list containing text (as strings) and image Parts (as types.Part).
    """
    print(f"Parsing '{xlsx_path}' for data and images...")
    
    # This list will hold our final, ordered content for the API
    content_parts = []

    # --- Step 1: Extract data tables using Pandas ---
    try:
        xls = pd.ExcelFile(xlsx_path)
    except Exception as e:
        print(f"Error reading file with pandas: {e}")
        return []

    for sheet_name in xls.sheet_names:
        print(f"  -> Processing Sheet: '{sheet_name}'")
        df = pd.read_excel(xls, sheet_name=sheet_name).dropna(how='all')
        
        if df.empty:
            continue
            
        # Add the sheet name as a header and the data as a Markdown table
        content_parts.append(f"--- Data from Sheet: {sheet_name} ---")
        content_parts.append(df.to_markdown(index=False))

    # --- Step 2: Extract images and their locations using openpyxl ---
    try:
        workbook = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        print(f"Error reading file with openpyxl: {e}")
        return content_parts # Return what we have so far

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if not sheet._images:
            continue

        print(f"  -> Extracting {len(sheet._images)} image(s) from '{sheet_name}'")
        for image in sheet._images:
            # Get the anchor cell (e.g., 'C2')
            row = image.anchor._from.row + 1
            col = image.anchor._from.col + 1
            cell_address = f"{get_column_letter(col)}{row}"
            
            # Get the image bytes
            image_data = image.image.fp.read()
            img_format = get_image_format(image_data)
            
            # This is the crucial part:
            # 1. Add a text label for the image.
            # 2. Add the image data itself as a Part.
            content_parts.append(f"\nImage anchored at cell {cell_address} in sheet '{sheet_name}':")
            content_parts.append(types.Part.from_bytes(
                data=image_data,
                mime_type=f'image/{img_format}'
            ))

    print("Parsing complete.")
    return content_parts